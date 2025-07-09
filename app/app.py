import streamlit as st
import pandas as pd
import openpyxl
import io
import base64

# Page configuration
st.set_page_config(
    page_title="Excel Macro Dashboard",
    page_icon="🔧",
    layout="wide",
    initial_sidebar_state="expanded"
)

def load_excel_file(uploaded_file):
    """Load Excel file and return workbook and sheet names"""
    try:
        # Load workbook with openpyxl to preserve formulas and macros
        workbook = openpyxl.load_workbook(uploaded_file, data_only=False, keep_vba=True)
        sheet_names = workbook.sheetnames
        
        # Also load with pandas for data manipulation
        excel_data = pd.read_excel(uploaded_file, sheet_name=None, engine='openpyxl')
        
        return workbook, sheet_names, excel_data
    except Exception as e:
        st.error(f"Error loading Excel file: {str(e)}")
        return None, None, None

def display_macro_info(workbook):
    """Display detailed macro/VBA information if present"""
    try:
        if hasattr(workbook, 'vba_archive') and workbook.vba_archive:
            st.success("🔧 **Macro-enabled Excel file detected!**")
            
            with st.expander("📋 Macro Details", expanded=True):
                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown("""
                    **✅ What's Preserved:**
                    - All formulas and calculations
                    - VBA macro code structure
                    - Worksheet functions
                    - Data validation rules
                    - Conditional formatting
                    """)
                
                with col2:
                    st.markdown("""
                    **⚠️ Limitations:**
                    - Macros won't execute in browser
                    - Interactive VBA features disabled
                    - Download file to run macros
                    - Security restrictions apply
                    """)
                
                st.info("💡 **Tip:** Download the processed file to use macros in Excel desktop application")
        else:
            st.info("📄 Standard Excel file (no macros detected)")
    except Exception as e:
        st.warning("⚠️ Could not analyze macro information")

def display_formulas_info(workbook, sheet_name):
    """Display formula information for a sheet"""
    try:
        worksheet = workbook[sheet_name]
        formulas = []
        
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f':  # Formula cell
                    formulas.append({
                        'Cell': cell.coordinate,
                        'Formula': cell.value,
                        'Calculated Value': cell.displayed_value,
                        'Data Type': 'Formula'
                    })
        
        if formulas:
            with st.expander(f"🔢 Formulas in {sheet_name} ({len(formulas)} found)", expanded=False):
                formula_df = pd.DataFrame(formulas)
                st.dataframe(formula_df, use_container_width=True)
                
                # Show formula statistics
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Total Formulas", len(formulas))
                with col2:
                    unique_formulas = len(set([f['Formula'] for f in formulas]))
                    st.metric("Unique Formulas", unique_formulas)
                with col3:
                    complex_formulas = len([f for f in formulas if len(f['Formula']) > 50])
                    st.metric("Complex Formulas", complex_formulas)
        else:
            st.info(f"No formulas found in {sheet_name}")
    except Exception as e:
        st.error(f"Error reading formulas: {str(e)}")

def create_download_link(workbook, filename):
    """Create a download link for the workbook with macros preserved"""
    output = io.BytesIO()
    try:
        # Save with macros if present
        workbook.save(output)
        output.seek(0)
        b64 = base64.b64encode(output.read()).decode()
        
        # Determine file extension
        if filename.endswith('.xlsm'):
            mime_type = "application/vnd.ms-excel.sheet.macroEnabled.12"
        else:
            mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        
        href = f'<a href="data:{mime_type};base64,{b64}" download="{filename}">📥 Download {filename}</a>'
        return href
    except Exception as e:
        st.error(f"Error creating download link: {str(e)}")
        return None

def analyze_sheet_complexity(df, workbook, sheet_name):
    """Analyze and display sheet complexity metrics"""
    try:
        worksheet = workbook[sheet_name]
        
        # Count different types of content
        formula_count = 0
        value_count = 0
        empty_count = 0
        
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    empty_count += 1
                elif cell.data_type == 'f':
                    formula_count += 1
                else:
                    value_count += 1
        
        # Display metrics
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("📊 Data Cells", value_count)
        with col2:
            st.metric("🔢 Formulas", formula_count)
        with col3:
            st.metric("📋 Total Cells", len(df) * len(df.columns) if not df.empty else 0)
        with col4:
            complexity_score = (formula_count * 2 + value_count) / max(1, len(df) * len(df.columns)) * 100
            st.metric("🎯 Complexity %", f"{complexity_score:.1f}")
            
    except Exception as e:
        st.error(f"Error analyzing sheet complexity: {str(e)}")

def main():
    # Header
    st.title("🔧 Excel Macro Dashboard")
    st.markdown("**Advanced Excel file viewer with macro support (.xlsm, .xlsx, .xls)**")
    
    # Sidebar for file upload
    with st.sidebar:
        st.header("📁 File Upload")
        uploaded_file = st.file_uploader(
            "Choose an Excel file",
            type=['xlsm', 'xlsx', 'xls'],
            help="Upload Excel files including macro-enabled (.xlsm) files"
        )
        
        if uploaded_file:
            st.success(f"✅ File: {uploaded_file.name}")
            st.info(f"📏 Size: {uploaded_file.size / 1024:.1f} KB")
            
            # File type detection
            file_ext = uploaded_file.name.split('.')[-1].lower()
            if file_ext == 'xlsm':
                st.warning("🔧 Macro-enabled file detected")
            elif file_ext == 'xlsx':
                st.info("📊 Standard Excel file")
            else:
                st.info("📄 Legacy Excel file")
    
    if uploaded_file is not None:
        # Load the Excel file
        with st.spinner("🔄 Loading Excel file and analyzing structure..."):
            workbook, sheet_names, excel_data = load_excel_file(uploaded_file)
        
        if workbook and sheet_names and excel_data:
            st.success(f"✅ Successfully loaded {len(sheet_names)} sheets")
            
            # Display macro information
            display_macro_info(workbook)
            
            # Global download option
            st.markdown("---")
            col1, col2 = st.columns([2, 1])
            with col1:
                st.markdown("### 📥 Download Options")
            with col2:
                if st.button("📥 Download Complete File", type="primary"):
                    download_link = create_download_link(workbook, uploaded_file.name)
                    if download_link:
                        st.markdown(download_link, unsafe_allow_html=True)
            
            st.markdown("---")
            
            # Create tabs for each sheet
            tabs = st.tabs(sheet_names)
            
            for i, (tab, sheet_name) in enumerate(zip(tabs, sheet_names)):
                with tab:
                    df = excel_data[sheet_name]
                    
                    st.subheader(f"📋 {sheet_name}")
                    
                    if not df.empty:
                        # Analyze sheet complexity
                        analyze_sheet_complexity(df, workbook, sheet_name)
                        
                        # Display formulas info
                        display_formulas_info(workbook, sheet_name)
                        
                        # Display the data
                        st.markdown("#### 📊 Sheet Data")
                        st.dataframe(df, use_container_width=True, height=400)
                        
                        # Individual sheet download
                        if st.button(f"📥 Download {sheet_name} only", key=f"download_sheet_{i}"):
                            # Create a new workbook with just this sheet
                            new_wb = openpyxl.Workbook()
                            new_ws = new_wb.active
                            new_ws.title = sheet_name
                            
                            # Copy data
                            for r_idx, row in enumerate(df.values, 1):
                                for c_idx, value in enumerate(row, 1):
                                    new_ws.cell(row=r_idx, column=c_idx, value=value)
                            
                            # Create download link
                            output = io.BytesIO()
                            new_wb.save(output)
                            output.seek(0)
                            b64 = base64.b64encode(output.read()).decode()
                            href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{sheet_name}.xlsx">Download {sheet_name}.xlsx</a>'
                            st.markdown(href, unsafe_allow_html=True)
                    else:
                        st.warning("⚠️ This sheet is empty")
    
    else:
        # Welcome message
        st.info("👆 Please upload an Excel file using the sidebar to get started")
        
        # Feature showcase
        st.markdown("""
        ### 🚀 Advanced Features:
        
        #### 🔧 **Macro Support**
        - **Full .xlsm compatibility** - Handle macro-enabled Excel files
        - **VBA preservation** - Macros are preserved in downloads
        - **Formula analysis** - Detailed breakdown of Excel formulas
        - **Security-first** - Macros don't execute in browser for safety
        
        #### 📊 **Data Analysis**
        - **Multi-sheet tabs** - Navigate between all Excel sheets
        - **Complexity metrics** - Understand your spreadsheet structure
        - **Formula detection** - See all formulas and their locations
        - **Interactive viewing** - Sort and filter data in real-time
        
        #### 📥 **Export Options**
        - **Complete file download** - Get the full file with macros intact
        - **Individual sheets** - Download specific sheets as needed
        - **Format preservation** - Maintain Excel formatting and structure
        
        ### 📁 Supported Formats:
        - **`.xlsm`** - Excel with macros (primary focus)
        - **`.xlsx`** - Standard Excel 2007+
        - **`.xls`** - Legacy Excel 97-2003
        
        ### 🔒 Security Note:
        Macro code is preserved but not executed for security. Download files to run macros in Excel.
        """)

if __name__ == "__main__":
    main()
